#!/usr/bin/python3
# -*- coding: UTF-8 -*-

import os, sys
import time
import calendar

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

import tkinter as tk
from tkinter.filedialog import askopenfilename
from tkinter.scrolledtext import ScrolledText

class ProjectMemenbers:
    """
    项目信息表
    """

    def __init__(self, xlsx, log=None, checked=False):
        self.xlsx = xlsx
        self.projects = {}
        self.members = {}
        self.log = log
        self.checked = checked

    def valid_date(self, start, end):
        if len(start) != 10 or len(end) != 10:
            self.log.error(f"时间格式错误，开始时间{start},结束时间{end}，正确的时间格式为YYYY-MM-DD")
            return False

        if start > end:
            self.log.error(f"时间计划错误，开始时间{start}大于结束时间{end}")
            return False
        return True

    def valid_parser(self, row):
        project_name = row[0].value
        if project_name is None:
            self.log.error("表格中存在二级项目名称为空！")
            return False

        project_start = row[1].value
        if project_start is None:
            self.log.error(f"项目{project_name}的起始时间不能为空！")
            return False

        project_end = row[2].value
        if project_end is None:
            project_end = time.strftime("%Y-%m-%d", time.localtime())

        if not self.valid_date(project_start, project_end):
            return False

        project_weight = row[3].value
        if project_weight is None:
            project_weight = 1
        else:
            if not isinstance(project_weight, int):
                self.log.warn(f"{project_weight}必须为数字！")
                return False

        namelist = row[4].value
        self.projects[project_name.strip()] = {'start': project_start.strip(),
                                               'end': project_end.strip(),
                                               'weight': project_weight}

        return self.insert_member(project_name, namelist)

    def insert_member(self, project, namelist):
        namelist = namelist.replace("，", ",").replace("\n", "")
        if namelist[-1] == ',':
            namelist = namelist[0:-1]
        for name in namelist.replace("，", ",").split(","):
            if len(name) < 8:
                self.log.error(f"[{project}]项目成员工号{name}不能小于8位")
                return False
            uid = name.strip()[:8]
            if uid in self.members:
                if project in self.members.get(uid):
                    continue
                self.members[uid].append(project)
            else:
                self.members[uid] = [project]
        return True

    def get_time_scale(self, sid, sname, cost, record_projects, year, month, date):
        """
        :param sid: 员工工号
        :param sid: 员工姓名
        :param cost:  工时
        :param record_projects:  本月参与的项目
        :param year: 年份
        :param month: 月份
        :param date: 某日
        :return: [ project_cost, project_cost2, project_cost3 ] 在每个项目上分配的工时
        """

        # 当日未到岗，或者尚未加入项目，工时为0
        if cost is None or int(cost) == 0:
            return [c * 0.00 for c in range(0, len(record_projects))]

        if not isinstance(cost, int) and not isinstance(cost, float):
            if self.checked:
                self.log.warn(f"员工 {sid}{sname} 在{year}-{month}-{date}工时数据{cost}必须为整形或浮点型！")
            return [c * 0.00 for c in range(0, len(record_projects))]

        # 判断某日是否有参与项目
        sum_weight = 0  # 计算分母
        work_projects = []  # 当日填写工时的项目清单, [项目名称,工时权重,分配工时值（保留小数点后两位）]
        date_fmt = f"{year}-{month}-{date}"
        for proj in record_projects:
            if date_fmt < self.projects[proj[0]]['start'] or date_fmt > self.projects[proj[0]]['end']:
                continue  # 当日未加入到项目工作，不分配工时
            sum_weight = sum_weight + self.projects[proj[0]]['weight']
            work_projects.append([proj[0], self.projects[proj[0]]['weight'], 0.00])

        # 当日为项目空窗期，相当于当日没有任何项目可以落工时
        if len(work_projects) <= 0:
            if self.checked:
                self.log.warn(f"员工 {sid}{sname} 在{year}-{month}-{date}处于项目空档期，工时无法落入项目！")
            return [c * 0.00 for c in range(0, len(record_projects))]

        # 计算分子（每个项目分配的工时时长，保留小数点后两位）
        # judge_value 修正由于未除尽导致精度缺失，列表中最后一个项目工时 = cost - 前面项目工时之和
        judge_value = 0.00
        for project_idx, project_value in enumerate(work_projects):
            if project_idx == len(work_projects) - 1:
                project_value[2] = round(cost - judge_value, 2)
            else:
                project_value[2] = round((project_value[1] / sum_weight) * cost, 2)
                judge_value = judge_value + project_value[2]

        # 返回所有需要记录项目当日的工时时值
        # record_weight = []
        tmpdict = {v[0]: v[2] for v in work_projects}
        for rec in record_projects:
            if rec[0] in tmpdict.keys():
                rec[1] = tmpdict[rec[0]]
            else:
                rec[1] = 0.00

        return [v[1] for v in record_projects]

    def get_month_projects(self, sid, sname, year, month):
        if sid not in self.members.keys():
            if self.checked:
                self.log.warn(f"员工 {sid}{sname} 在项目成员信息表中未查询到相关记录！")
            return None

        record_projects = []
        month_fmt = f"{year}-{month}"
        for name in self.members[sid]:
            # 当月任何一天在项目中记录了工时，那么全月该项目都要记录工时，如在8月31日时记录A项目工时
            # 那么本月需要增加全月A项目的工时记录，除8月31日之外的日期记录为0，只是为了保持记录整齐
            if self.projects[name]['start'][:7] <= month_fmt <= self.projects[name]['end'][:7]:
                record_projects.append([name, 0])

        return record_projects

    def print(self):
        for k, v in self.members.items():
            print(k, v)

        for k, v in self.projects.items():
            print(k, v)

    def parser(self):
        self.log.info(f"正在读取项目成员信息表[{self.xlsx}] ...")
        wb = load_workbook(self.xlsx, read_only=True)
        ws = wb.active
        if ws.title != "项目及成员管理":
            self.log.error(f"导入的的项目成员信息表{self.xlsx}数据内容错误，请确认是否导入了正确的表？")
            return False
        for row in ws.iter_rows(min_row=3):
            # 如果项目没有成员，则忽略
            if row[4].value is None:
                continue
            if not self.valid_parser(row):
                return False

        self.log.info(f"项目成员信息数据读取成功。")
        # self.print()
        return True


class DataProduct:
    """
    读取生产数据
    """

    def __init__(self, xlsx, project, log=None, checked=False):
        self.xlsx = xlsx
        self.data = {}
        self.project = project
        self.log = log
        self.checked = checked

    def time_analysis(self, year, month, record):
        """
        工时数据分析
        """

        # 1. 获取员工本月所参与的项目，根据项目权重计算工时比例，每个项目新增一行数据记录工时
        record_projects = self.project.get_month_projects(record[0], record[1], year, month)
        if record_projects is None:
            return None

        info = record[:6]  # 员工基本信息
        info.append([v[0] for v in record_projects])  # 当月参加项目列表, 此项为info[6]
        info.append([c * 0.00 for c in range(0, len(record_projects))])  # 当月工时汇总, 此项为info[7]

        for idx in range(1, len(record[7:]) + 1):
            time_cost = self.project.get_time_scale(record[0], record[1],
                                                    record[idx + 6],  # 从表格第7列（列表下标为6），也就是1日开始
                                                    record_projects,
                                                    year, month, "{0:>02d}".format(idx))  # 进行汇总
            info[7] = [round(i + j, 2) for i, j in zip(info[7], time_cost)]
            # info[7] = numpy.array(copy.deepcopy(info[7])) + numpy.array(time_cost)
            info.append(time_cost)
        return info

    def parser(self):
        filename = os.path.basename(self.xlsx)
        yearname = filename.split("-")[1][:4]
        if yearname != "2020" and yearname != "2021":
            self.log.error(f"工时数据表[{self.xlsx}]文件名错误\n\t'文档名称-'后至少要有4位数字年份，如'产研平台工时数据-202101-08(XXX).xlsx'")
            return False

        self.data[yearname] = {}
        wb = load_workbook(self.xlsx, read_only=True)

        self.log.info(f"读取工时数据信息表数据 ...")
        month_data = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if "月" not in sheet_name:
                self.log.error(f"传入的工时数据表{self.xlsx}数据内容错误，请确认是否导入了正确的表？")
                return False
            month_num = sheet_name.replace("月", "")
            month_num = "{0:>02s}".format(month_num)
            month_data[sheet_name] = []

            self.log.info(f"读取{yearname}-{sheet_name}数据 ...")
            # 计算某有多少天，用于统计列数，防止表格数据列之外存在垃圾数据读取。
            _, numbers = calendar.monthrange(int(yearname), int(month_num))
            # 两者取小的，假设某月并没有完成统计全部天数，或者表格存在垃圾列
            max_column = min(numbers + 7, ws.max_column)

            r = 0
            for row in ws.rows:
                record = [row[c].value for c in range(0, max_column)]
                if record[0] is None or not isinstance(record[0], str) or len(record[0]) < 0:
                    if self.checked:
                        self.log.warn(f"SheetName[{sheet_name}]存在空行[行号：{r + 1}]或者数据不完整的行\n{record}")
                    continue
                if r == 0:
                    record.insert(6, '项目')
                    month_data[sheet_name].append(record)
                else:
                    tmp_record = self.time_analysis(yearname, month_num, record)
                    if tmp_record is not None:
                        month_data[sheet_name].append(tmp_record)
                r = r + 1
        self.log.info(f"工时数据信息表数据读取完成。")
        self.data[yearname] = month_data
        wb.close()

        return True

    def writer(self, filename):
        self.log.info(f"正在处理工时数据 ...")
        wb = Workbook(write_only=True)
        for month_data in self.data.values():
            for sheetname, records in month_data.items():
                ws = wb.create_sheet(title=sheetname)
                # rows = len(records)
                # columns = len(records[0])
                for i, rec in enumerate(records):
                    if i == 0:
                        # 设置列宽
                        for col_id in range(9, len(rec) + 1):
                            col_letter = get_column_letter(col_id)
                            ws.column_dimensions[col_letter].width = 5
                        ws.append(rec)
                    else:
                        # ws.append(rec)
                        revert_list = list(map(list, zip(*rec[6:])))
                        for line in revert_list:
                            ws.append(rec[:6] + line)
                    # print(i, rec)

        wb.save(filename)
        wb.close()

        self.log.info(f"工时数据处理完成，生成新表：{filename} ...")
        return True


class DataProcess:
    """
    后台程序独立统一入口
    """

    def __init__(self, log, data_file, project_file):
        self.log = log
        self.data_file = data_file
        self.project_file = project_file

    def run(self, checked=False):

        try:
            self.log.info(f"\n**** 开始进行工时数据解析处理 ...")

            # 解析项目成员信息表
            pminfo = ProjectMemenbers(self.project_file, self.log, checked)
            if not pminfo.parser():
                return

            # 解析处理工时信息表
            filename = os.path.basename(self.data_file)
            outfile = os.path.join(os.getcwd(), f"TimeResults_{filename}")
            productor = DataProduct(self.data_file, pminfo, self.log, checked)
            if not productor.parser():
                return

            # 转换生成新的数据
            if not productor.writer(outfile):
                return

        except Exception as e:
            self.log.error(str(e))
            raise e


class LogTrace:
    """
    日志打印
    """

    def __init__(self):
        self.log_gui = None
        self.log_file = None
        self.inf = None

    def init_log(self, log_gui=None, log_file=None):
        self.log_gui = log_gui
        self.log_file = log_file
        if self.log_file is not None:
            self.inf = open(self.log_file, 'a+')

    def log(self, level, msg):
        if self.log_gui is not None:
            self.log_gui.config(state=tk.NORMAL)
            self.log_gui.see(tk.END)

            if level == "ERROR":
                self.log_gui.insert(tk.INSERT, f"[{level}] {msg}\n")
            elif level == "WARN":
                self.log_gui.insert(tk.INSERT, f"[{level}] {msg}\n")
            else:
                self.log_gui.insert(tk.INSERT, f"[{level}] {msg}\n")
            self.log_gui.update()

            self.log_gui.see(tk.END)
            self.log_gui.config(state=tk.DISABLED)

        if self.inf is not None:
            self.inf.write(f"[{level}] {msg}\n")

        # if self.log_gui is None and self.log_file is None:
        print(f"[{level}] {msg}")

    def info(self, msg):
        self.log("INFO", msg)

    def warn(self, msg):
        self.log("WARN", msg)

    def error(self, msg):
        self.log("ERROR", msg)


class EasyGui:
    """
    简易操作界面
    """

    def __init__(self):
        self.win = None
        self.scroll = None  # 滚动文本框

        # 收录原始数据文件
        self.data_file = ""
        self.project_file = ""
        self.check_option = False  # 默认不开启数据检验
        self.log = LogTrace()  # 处理信息文本输出框

    def use_help(self):
        self.scroll.config(state=tk.NORMAL)
        self.scroll.see(tk.END)
        self.scroll.insert(tk.INSERT, '''
==============================================================================================
1.请先选择工时数据信息表，请注意'文档名称-'后至少要有4位数字年份，如'产研平台工时数据-202101-08(XXX).xlsx'。
2.请注意当前仅支持word 2007，即文件格式为xlsx的Excel文件，如您是从石墨文档中导出的文件，有可能不能直接读取，
  可尝试本地另存为一份xlsx格式文件。
2.为了提高数据转换效率，请尽量保障原始文件中不包含垃圾数据，比如与工时数据无关的多余空行，空列，空字段等。 
3.开启数据检查后，可能会出现大量的提示或者警告信息，请尝试修复，或者确认是无用数据后关掉检查。警告信息不影响
  数据转换。
4.转换后的文件在当前目录下以<TimeResults_原文件名>命名。
    
                                                                欢迎测试Bug，任何问题请联系封立刚。
                                                                                2021年9月19日
==============================================================================================
                                                      \n''')
        self.scroll.update()
        self.scroll.see(tk.END)
        self.scroll.config(state=tk.DISABLED)

    # 清空日志
    def clear_log(self):
        self.scroll.config(state=tk.NORMAL)

        self.scroll.delete(0.0, tk.END)
        self.scroll.update()

        self.scroll.config(state=tk.DISABLED)

    def process(self, btn):
        # 初始化日志处理对象
        self.log.init_log(self.scroll, None)

        if len(self.data_file) <= 0:
            self.log.error(f'您还未导入工时数据信息表.')
            self.log.info(f'这是个是正常的显示 ')
            return
        if len(self.project_file) <= 0:
            self.log.error(f'您还未导入项目成员信息表.')
            return

        try:
            # 初始化日志处理对象
            self.log.init_log(self.scroll, None)

            # 禁用处理按钮，避免重复进入
            btn.config(state='disabled')

            DataProcess(self.log,
                        self.data_file,
                        self.project_file).run(self.check_option)
        except Exception:
            raise
        finally:
            btn.config(state='normal')

    def check_selection(self, checkval):
        self.check_option = checkval.get()

    def openfile(self, snames, opt):
        file_name = askopenfilename(title='选择Excel文件',
                                    initialdir=os.getcwd(),
                                    multiple=False,
                                    filetypes=[('Excel', '*.xlsx'), ('All Files', '*')])
        print(file_name)
        snames.set(file_name)
        if opt == 1:
            self.data_file = file_name
        else:
            self.project_file = file_name
        return file_name

    def file_dialog(self, frm, frm_width, frm_height):
        data_notes = ['请导入工时数据(仅支持xlsx格式数据）...']
        project_notes = ['请导入项目成员信息(仅支持xlsx格式数据）...']
        data_names = tk.StringVar()
        project_names = tk.StringVar()
        data_names.set(data_notes)
        project_names.set(project_notes)

        # 标签控件
        tk.Label(frm, text="工时数据：", width=10, font=("宋体", 11)).grid(row=0, sticky=tk.E)  # 靠右
        tk.Label(frm, text="项目数据：", width=10, font=("宋体", 11)).grid(row=1, sticky=tk.E)  # 第二行，靠左

        # 输入控件
        tk.Entry(frm, textvariable=data_names).grid(row=0, column=1, ipadx=200, ipady=2)
        tk.Entry(frm, textvariable=project_names).grid(row=1, column=1, ipadx=200, ipady=2)

        # 按钮控件
        tk.Button(frm, text="选择", command=lambda: self.openfile(data_names, 1),
                  width=10).grid(row=0, column=2, pady=1)
        tk.Button(frm, text="选择", command=lambda: self.openfile(project_names, 2),
                  width=10).grid(row=1, column=2, pady=1)

        # 其它
        checkvar = tk.BooleanVar()
        tk.Checkbutton(frm, text='开启数据检查', variable=checkvar, onvalue=True, offvalue=False,
                       command=lambda: self.check_selection(checkvar)).grid(row=2, column=1, sticky=tk.W)
        tk.Button(frm, text='帮助', width=10, command=self.use_help).grid(row=2, column=1, sticky=tk.E)
        btn = tk.Button(frm, text='执行', width=10, command=lambda: self.process(btn))
        btn.grid(row=2, column=2, pady=5, sticky=tk.W)

    def text_dialog(self, frm, frm_width, frm_height):
        self.scroll = ScrolledText(frm, width=frm_width, height=frm_height, wrap=tk.WORD,state=tk.DISABLED )
        self.scroll.pack(fill=tk.X, ipady=2, expand=False)

    def init_window(self):
        window = tk.Tk()
        window.title('快乐工时')

        width, height = 1000, 600
        screenwidth = window.winfo_screenwidth()
        screenheight = window.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        window.geometry(alignstr)
        window['bg'] = '#BCD2EE'
        #window.attributes('-alpha', 100)
        window['borderwidth'] = 3
        # 设置窗口大小不可改变
        #window.resizable(width=False, height=False)
        window.update()
        return window

    def application(self):
        win = self.init_window()  # 初始化窗口位置，大小

        win_width = win.winfo_width()
        win_height = win.winfo_height()

        frm1_height = 150
        frm1 = tk.Frame(master=win, relief='solid',
                        bd=1, height=frm1_height, width=win_width, padx=2, pady=3)
        frm1.pack(side=tk.TOP)
        frm1.update()

        # 边框的3D样式 flat、sunken、raised、groove、ridge、solid。
        frm2 = tk.Frame(master=win, relief='ridge', bd=3,
                        height=win_height - frm1_height, width=win_width, padx=1, pady=1)
        frm2.pack(side=tk.BOTTOM)
        frm2.update()

        self.file_dialog(frm1, frm1.winfo_width(), frm1.winfo_height())
        self.text_dialog(frm2, frm2.winfo_width(), frm2.winfo_height())
        win.update()

        win.mainloop()


def app_main():
    EasyGui().application()


def cmd_main(argv):
    if len(argv) <= 3:
        print(f" ERR: 命令参数错误。")
        print(f" 示例: {argv[0]} <工时数据信息表.xlsx> <项目成员信息表.xlsx> [True|False]")
        return False
    log = LogTrace()
    data_file = argv[1]
    if not os.path.isfile(data_file):
        if not os.path.isfile(os.path.join(os.getcwd(), data_file)):
            print(f"找不到工时数据表文件: {data_file}")
            return False
    project_file = argv[2]
    if not os.path.isfile(project_file):
        if not os.path.isfile(os.path.join(os.getcwd(), project_file)):
            print(f"找不到成员项目信息表文件: {project_file}")
            return False

    checked = False
    if len(argv) == 4:
        if argv[3].upper() == "TRUE":
            checked = True
        elif argv[3].upper() == "FALSE":
            checked = False
        else:
            print(f" ERR: 命令参数错误。")
            print(f" 示例: {argv[0]} <工时数据信息表.xlsx> <项目成员信息表.xlsx> [True|False]")

    DataProcess(log, data_file, project_file).run(checked)


if '__main__' == __name__:
    # cmd_main(sys.argv)
    app_main()
